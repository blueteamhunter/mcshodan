1) High-level architecture

Account A (tooling): hosts the Lambda, SES (email), and an EventBridge schedule. Optional S3 bucket for overflow (large attachments).

Accounts B & C (targets): expose a read-only cross-account role that Account A can assume to list Route 53 zones and records.

Flow: EventBridge (cron) → Lambda in A → STS AssumeRole into B & C → list zones & all records → create one Excel per zone → email via SES (or upload ZIP(s) to S3 and email signed links if needed).

2) SES setup (Account A)

Pick SES region (e.g., us-east-1).

Verify sender domain or From address (configure SPF/DKIM).

If SES is in sandbox, either:

Move to production, or

Verify all recipient addresses.

Keep:

FROM_ADDRESS (e.g., ops@yourdomain.com)

Recipients list for TO_ADDRESSES.

3) Cross-account roles in Accounts B & C

Create the same role (name idea: Route53ReadExport) in B and C.

Trust policy (replace ACCOUNT_A_ID)
{
  "Version": "2012-10-17",
  "Statement": [
    {
      "Sid": "TrustAccountA",
      "Effect": "Allow",
      "Principal": { "AWS": "arn:aws:iam::ACCOUNT_A_ID:root" },
      "Action": "sts:AssumeRole"
    }
  ]
}

Permissions policy

Use the AWS managed AmazonRoute53ReadOnlyAccess or this minimal inline:

{
  "Version": "2012-10-17",
  "Statement": [
    { "Effect": "Allow",
      "Action": ["route53:ListHostedZones","route53:ListResourceRecordSets"],
      "Resource": "*" }
  ]
}


Record the ARNs, e.g.:

arn:aws:iam::ACCOUNT_B_ID:role/Route53ReadExport

arn:aws:iam::ACCOUNT_C_ID:role/Route53ReadExport

4) Lambda execution role (Account A)

Create an IAM role for the Lambda (e.g., r53-exporter-lambda-role).

Permissions:

{
  "Version": "2012-10-17",
  "Statement": [
    { "Sid": "Logs", "Effect": "Allow",
      "Action": ["logs:CreateLogGroup","logs:CreateLogStream","logs:PutLogEvents"],
      "Resource": "arn:aws:logs:*:*:*" },

    { "Sid": "AssumeTargets", "Effect": "Allow",
      "Action": ["sts:AssumeRole"],
      "Resource": [
        "arn:aws:iam::ACCOUNT_B_ID:role/Route53ReadExport",
        "arn:aws:iam::ACCOUNT_C_ID:role/Route53ReadExport"
      ]
    },

    { "Sid": "SES", "Effect": "Allow",
      "Action": ["ses:SendRawEmail"],
      "Resource": "*" },

    { "Sid": "S3Optional", "Effect": "Allow",
      "Action": ["s3:PutObject","s3:GetObject","s3:ListBucket"],
      "Resource": [
        "arn:aws:s3:::YOUR_EXPORT_BUCKET",
        "arn:aws:s3:::YOUR_EXPORT_BUCKET/*"
      ]
    }
  ]
}


If your bucket uses KMS, also grant the Lambda role kms:Encrypt/kms:Decrypt on the key.

5) Lambda layer (openpyxl)

We’ll write .xlsx via openpyxl (pure Python). Build the layer on a Linux env matching Lambda runtime:

mkdir -p layer/python
pip install --upgrade pip
pip install --target layer/python openpyxl==3.1.5
cd layer && zip -r ../openpyxl-layer.zip python && cd ..


Create a Lambda Layer from openpyxl-layer.zip. Keep the ARN.

6) Lambda configuration

Runtime: Python 3.12

Memory: 512–1024 MB

Timeout: 5–10 minutes (increase if many zones/records)

Layers: attach the openpyxl layer

Environment variables:

TARGET_ROLE_ARNS = arn:aws:iam::B:role/Route53ReadExport,arn:aws:iam::C:role/Route53ReadExport

SES_REGION = us-east-1

FROM_ADDRESS = ops@yourdomain.com

TO_ADDRESSES = you@yourdomain.com,security@yourdomain.com

ATTACHMENT_STRATEGY = auto (attach | link | auto)

SIZE_LIMIT_MB = 7 (keep raw email < ~10MB)

MAX_ATTACHMENTS = 10 (cap to avoid email client issues)

ZIP_STRATEGY = per_account (none | per_account | all)

S3_BUCKET = your-export-bucket (required when links are used)

PRESIGN_TTL_SEC = 604800 (7 days)

7) Lambda code (one Excel per zone, emailed via SES; auto-fallback to S3 links)

Drop this as lambda_function.py. It includes the fixed SES helper and the change to one workbook per zone.

import os, io, json, zipfile, time, datetime, logging, re
import boto3
from botocore.config import Config
from email.message import EmailMessage
from email.utils import formatdate, make_msgid
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

log = logging.getLogger()
log.setLevel(logging.INFO)

ROUTE53_REGION = "us-east-1"  # Route 53 is Global; us-east-1 works for boto3 client
TMP_DIR = "/tmp"

def _env(name, default=None, required=False):
    val = os.environ.get(name, default)
    if required and not val:
        raise RuntimeError(f"Missing required env var: {name}")
    return val

def assume_session(role_arn: str, session_name: str, duration=3600) -> boto3.Session:
    sts = boto3.client("sts")
    creds = sts.assume_role(RoleArn=role_arn, RoleSessionName=session_name, DurationSeconds=duration)["Credentials"]
    return boto3.Session(
        aws_access_key_id=creds["AccessKeyId"],
        aws_secret_access_key=creds["SecretAccessKey"],
        aws_session_token=creds["SessionToken"],
    )

def list_all_zones(r53_client):
    zones, marker = [], None
    while True:
        resp = r53_client.list_hosted_zones(Marker=marker) if marker else r53_client.list_hosted_zones()
        zones.extend(resp.get("HostedZones", []))
        if resp.get("IsTruncated"):
            marker = resp.get("NextMarker")
        else:
            break
    return zones

def list_all_records(r53_client, zone_id):
    records, params = [], {"HostedZoneId": zone_id}
    while True:
        resp = r53_client.list_resource_record_sets(**params)
        records.extend(resp.get("ResourceRecordSets", []))
        if resp.get("IsTruncated"):
            params["StartRecordName"] = resp["NextRecordName"]
            params["StartRecordType"] = resp["NextRecordType"]
            if "NextRecordIdentifier" in resp:
                params["StartRecordIdentifier"] = resp["NextRecordIdentifier"]
        else:
            break
    return records

def sanitize_sheet_name(name: str) -> str:
    # Excel sheet name max 31 chars; disallow : \ / ? * [ ]
    name = re.sub(r'[:\\/\?\*\[\]]', '-', name)
    return name[:31] if len(name) > 31 else name

def sanitize_filename(name: str, maxlen=120) -> str:
    # safe for file path
    name = re.sub(r'[^A-Za-z0-9._-]+', '_', name)
    return name[:maxlen]

def make_workbook_for_zone(account_id: str, zone_name: str, zone_id: str, records: list) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = sanitize_sheet_name(zone_name or zone_id)

    headers = [
        "ZoneName","ZoneId","RecordName","Type","TTL","Values",
        "AliasDNSName","AliasHostedZoneId","EvaluateTargetHealth",
        "SetIdentifier","Weight","Failover","Region","MultiValueAnswer"
    ]
    ws.append(headers)

    for rr in records:
        name = rr.get("Name","")
        rtype = rr.get("Type","")
        ttl = rr.get("TTL","")  # might be absent for Alias
        values = "\n".join([v.get("Value","") for v in rr.get("ResourceRecords", [])])

        alias = rr.get("AliasTarget", {})
        row = [
            zone_name, zone_id, name, rtype, ttl, values,
            alias.get("DNSName",""), alias.get("HostedZoneId",""), alias.get("EvaluateTargetHealth",""),
            rr.get("SetIdentifier",""), rr.get("Weight",""), rr.get("Failover",""),
            rr.get("Region",""), rr.get("MultiValueAnswer","")
        ]
        ws.append(row)

    # Autosize columns (bounded to keep sheets readable)
    for col_idx, _ in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            val = str(cell.value) if cell.value is not None else ""
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    ts = datetime.datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    safe_zone = sanitize_filename(zone_name.rstrip(".") if zone_name else zone_id)
    path = f"{TMP_DIR}/route53-{account_id}-{safe_zone}-{ts}.xlsx"
    wb.save(path)
    return path

def zip_files(filepaths, zip_name=None):
    if not zip_name:
        ts = datetime.datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
        zip_name = f"{TMP_DIR}/route53-exports-{ts}.zip"
    with zipfile.ZipFile(zip_name, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for fp in filepaths:
            zf.write(fp, arcname=os.path.basename(fp))
    return zip_name

def size_mb(path): return os.path.getsize(path) / (1024*1024)

def upload_to_s3(bucket, path, key=None, kms_key_id=None):
    s3 = boto3.client("s3")
    if not key:
        key = f"route53-exports/{os.path.basename(path)}"
    extra = {}
    if kms_key_id:
        extra["ServerSideEncryption"] = "aws:kms"
        extra["SSEKMSKeyId"] = kms_key_id
    s3.upload_file(path, bucket, key, ExtraArgs=extra)
    return key

def presign(bucket, key, expires=3600):
    s3 = boto3.client("s3")
    return s3.generate_presigned_url("get_object", Params={"Bucket": bucket, "Key": key}, ExpiresIn=expires)

def send_email_with_attachments(from_addr, to_addrs, subject, body_text, attachments):
    """
    attachments: list of { "filename": str, "content": bytes, "mime": "<type/subtype>" }
    """
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = from_addr
    msg["To"] = ", ".join(to_addrs)
    msg["Date"] = formatdate(localtime=True)
    msg["Message-Id"] = make_msgid()
    msg.set_content(body_text)

    for att in attachments:
        maintype, subtype = att["mime"].split("/", 1)
        msg.add_attachment(att["content"], maintype=maintype, subtype=subtype, filename=att["filename"])

    ses = boto3.client("ses", region_name=os.environ["SES_REGION"])
    ses.send_raw_email(Source=from_addr, Destinations=to_addrs, RawMessage={"Data": msg.as_bytes()})

def lambda_handler(event, context):
    # Env vars
    target_role_arns = [a.strip() for a in _env("TARGET_ROLE_ARNS", required=True).split(",") if a.strip()]
    from_addr = _env("FROM_ADDRESS", required=True)
    to_addrs = [a.strip() for a in _env("TO_ADDRESSES", required=True).split(",") if a.strip()]
    attach_strategy = _env("ATTACHMENT_STRATEGY", "auto").lower()  # attach|link|auto
    size_limit_mb = float(_env("SIZE_LIMIT_MB", "7"))
    max_attachments = int(_env("MAX_ATTACHMENTS", "10"))
    zip_strategy = _env("ZIP_STRATEGY", "per_account").lower()     # none|per_account|all
    s3_bucket = _env("S3_BUCKET", "")
    presign_ttl = int(_env("PRESIGN_TTL_SEC", "604800"))

    generated_files = []             # all .xlsx files (one per zone)
    files_by_account = {}            # account_id -> [paths]

    # Crawl each account
    for role_arn in target_role_arns:
        account_id = role_arn.split(":")[4]
        files_by_account.setdefault(account_id, [])

        log.info(f"Processing account {account_id} via {role_arn}")
        sess = assume_session(role_arn, session_name=f"r53-export-{int(time.time())}")
        r53 = sess.client("route53", region_name=ROUTE53_REGION,
                          config=Config(retries={"max_attempts": 10, "mode": "adaptive"}))

        zones = list_all_zones(r53)
        log.info(f"Account {account_id} zones: {len(zones)}")

        for z in zones:
            zone_id = z["Id"].split("/")[-1]
            zone_name = z.get("Name", "").rstrip(".")
            records = list_all_records(r53, zone_id)
            xlsx_path = make_workbook_for_zone(account_id, zone_name or zone_id, zone_id, records)
            generated_files.append(xlsx_path)
            files_by_account[account_id].append(xlsx_path)

    # Decide delivery
    subject = f"Route 53 DNS Export – {datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M:%SZ')} UTC"
    body_lines = [
        "Hello,",
        "",
        "Attached are the Route 53 exports (one Excel file per hosted zone).",
        "If attachments are too large or too many, download links (S3 pre-signed) are provided below.",
        ""
    ]

    total_mb = sum(size_mb(p) for p in generated_files)
    attachments = []

    def attach_files(paths):
        for p in paths:
            with open(p, "rb") as f:
                data = f.read()
            attachments.append({
                "filename": os.path.basename(p),
                "content": data,
                "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            })

    def upload_and_link(paths, title_prefix=None):
        urls = []
        for p in paths:
            key = upload_to_s3(s3_bucket, p)
            url = presign(s3_bucket, key, presign_ttl)
            urls.append((os.path.basename(p), url))
        if urls:
            if title_prefix:
                body_lines.append(f"{title_prefix}")
            for name, url in urls:
                body_lines.append(f"- {name}: {url}")

    # Strategy: attach if under size & count limits; otherwise zip (per_account or all); otherwise link
    if attach_strategy == "attach" or (attach_strategy == "auto" and total_mb <= size_limit_mb and len(generated_files) <= max_attachments):
        attach_files(generated_files)

    else:
        if not s3_bucket and zip_strategy == "none":
            raise RuntimeError("Too many/large attachments; set S3_BUCKET or enable zipping.")

        if zip_strategy in ("per_account", "all"):
            # Build zips
            zip_paths = []
            if zip_strategy == "per_account":
                for acct, paths in files_by_account.items():
                    if not paths:
                        continue
                    ts = datetime.datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
                    zip_path = zip_files(paths, zip_name=f"{TMP_DIR}/route53-exports-{acct}-{ts}.zip")
                    zip_paths.append(zip_path)
            else:
                zip_paths.append(zip_files(generated_files))

            zip_total_mb = sum(size_mb(p) for p in zip_paths)
            if attach_strategy == "attach" or (attach_strategy == "auto" and zip_total_mb <= size_limit_mb and len(zip_paths) <= max_attachments):
                # Attach zips
                for p in zip_paths:
                    with open(p, "rb") as f:
                        data = f.read()
                    attachments.append({
                        "filename": os.path.basename(p),
                        "content": data,
                        "mime": "application/zip"
                    })
            else:
                # Link the zips from S3
                if not s3_bucket:
                    raise RuntimeError("Attachment overflow; S3_BUCKET required to send links.")
                for p in zip_paths:
                    key = upload_to_s3(s3_bucket, p)
                    url = presign(s3_bucket, key, presign_ttl)
                    body_lines.append(f"- {os.path.basename(p)}: {url}")

        elif zip_strategy == "none":
            if not s3_bucket:
                raise RuntimeError("Attachment overflow and ZIP disabled; S3_BUCKET required.")
            upload_and_link(generated_files, title_prefix="Per-zone files:")
        else:
            raise RuntimeError(f"Unknown ZIP_STRATEGY: {zip_strategy}")

    body_text = "\n".join(body_lines)
    send_email_with_attachments(
        from_addr=from_addr,
        to_addrs=to_addrs,
        subject=subject,
        body_text=body_text,
        attachments=attachments
    )

    return {
        "ok": True,
        "zones_files": [os.path.basename(p) for p in generated_files],
        "attached": [a["filename"] for a in attachments],
        "total_mb": total_mb
    }


What changed vs. before

One workbook per zone via make_workbook_for_zone(...).

SES helper fixed (send_email_with_attachments is correct and used).

Smarter delivery:

Attach if under size and count limits.

Else zip per account (configurable) and attach or link from S3.

Fully configurable via env vars.

8) EventBridge schedule (São Paulo → UTC)

To run monthly on the 1st at 09:00 São Paulo (UTC-3):

cron(0 12 1 * ? *)


Target = your Lambda. No input required.

9) Operations & security tips

Least privilege: Targets only need ListHostedZones and ListResourceRecordSets.

Logging: The code logs counts and filenames—no record contents (avoid leaking TXT values).

S3 retention: Apply lifecycle (e.g., delete route53-exports/ after 30 days).

Encryption: Enable SSE-KMS on the bucket and grant the Lambda role access to the key.

Audit: CloudTrail tracks STS AssumeRole into B & C.

10) Validation checklist

 SES From domain/address verified; production access (or recipients verified).

 Lambda role has sts:AssumeRole to B & C and ses:SendRawEmail.

 openpyxl layer attached; runtime Python 3.12.

 Env vars set (see Section 6).

 EventBridge cron created and enabled.

 Test invoke: email received with one .xlsx per zone attached, or S3 links if large.
